"""
 ***openpyxl***
 openpyxl模块是用于处理Microsoft Excel文件的第三方库,可以对Excel文件中的数据进行写入和读取
 1. load_workbook(filename)     打开已存在的表格，结果为工作簿对象
 2. workbook.sheetnames         工作簿对象的sheetnames属性，用于获取所有工作表的名称，结果为列表类型
 3. sheet.append(lst)           向工作表中添加一行数据，新数据接在工作表已有数据的后面
 4. workbook.save(excelname)    保存工作簿
 5. Workbook()                  创建新的工作簿对象
"""

import openpyxl
import module_c
html=module_c.get_html()
lst=module_c.parse_html(html)
#创建一个新的Excel
workbook = openpyxl.Workbook()
#在Excel中创建工作表
sheet=workbook.create_sheet('景区天气')
#向工作表中添加数据
for item in lst:
    sheet.append(item)
workbook.save('./attachments/景区天气.xlsx')

#——————————————————————————————————————————————————
#打开工作簿
workbook=openpyxl.load_workbook('./attachments/景区天气.xlsx')
#选择要操作的工作表
sheet=workbook['景区天气']
#表格数据是二维列表
lst=[]#存储行数据
for row in sheet.rows:
    sublst=[]
    for cell in row:
        sublst.append(cell.value)
    lst.append(sublst)
for item in lst:
    print(item)